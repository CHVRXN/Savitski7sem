import random
import sqlite3
import time
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl
import os
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook

def loginPage(logdata):
    sup.destroy()
    global login
    login = Tk()

    user_name = StringVar()
    password = StringVar()
    show_password = IntVar()  # Флаг для определения видимости пароля
    window_width = 720
    window_height = 440

    # Получение размеров экрана
    screen_width = login.winfo_screenwidth()
    screen_height = login.winfo_screenheight()

    # Вычисление позиции окна для центрирования
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # Установка размеров и позиции окна
    login.geometry(f"{window_width}x{window_height}+{x}+{y}")

    login_canvas = Canvas(login, width=720, height=440, bg="pink")
    login_canvas.pack()

    login_frame = Frame(login_canvas, bg="white")
    login_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    heading = Label(login_frame, text="Авторизация", fg="black", bg="white")
    heading.config(font=('calibri 40'))
    heading.place(relx=0.2, rely=0.1)

    def toggle_password_visibility():
        if show_password.get() == 1:
            pas.config(show="")
        else:
            pas.config(show="*")

    show_password_checkbox = Checkbutton(login_frame, text="Показать пароль", variable=show_password,
                                         command=toggle_password_visibility)
    show_password_checkbox.place(relx=0.5, rely=0.6, anchor="center")  # Центрирование кнопки по горизонтали и вертикали

    # USER NAME
    ulabel = Label(login_frame, text="Имя пользователя", fg='black', bg='white')
    ulabel.place(relx=0.12, rely=0.4)
    uname = Entry(login_frame, bg='#d3d3d3', fg='black', textvariable=user_name)
    uname.config(width=42)
    uname.place(relx=0.31, rely=0.4)

    # PASSWORD
    plabel = Label(login_frame, text="Пароль", fg='black', bg='white')
    plabel.place(relx=0.215, rely=0.5)
    pas = Entry(login_frame, bg='#d3d3d3', fg='black', show="*", textvariable=password)
    pas.config(width=42)
    pas.place(relx=0.31, rely=0.5)

    def check():
        global global_username
        global testName
        for a, b, c, d in logdata:
            if b == uname.get() and c == pas.get():
                global_username = uname.get()
                messagebox.showinfo("Успешный вход", "Вы успешно вошли в систему!")
                if uname.get() == "admin" and pas.get() == "adminadmin":
                    adminpage()
                else:
                    menu()
                    # Отображение сообщения об успешном входе

                break
        else:
            error = Label(login_frame, text="Не правильное имя пользователя или пароль!", fg='black', bg='white')
            error.place(relx=0.37, rely=0.7)

    # LOGIN BUTTON
    log = Button(login_frame, text='Войти', padx=5, pady=5, width=4, command=check)
    log.configure(width=15, height=1, activebackground="#33B5E5", relief=FLAT)
    log.place(relx=0.4, rely=0.7)  # Увеличьте значение rely для смещения кнопки вниз

    login.mainloop()


def adminpage():
    global global_username

    def search_records():
        search_name = search_entry.get().strip()  # Получаем введенное имя пользователя
        selected_quiz = quiz_combobox.get()  # Получаем выбранное название викторины
        tree.delete(*tree.get_children())  # Очищаем Treeview перед отображением результатов поиска

        # Если поле для поиска пустое, выводим все данные из таблицы
        if not search_name:
            conn = sqlite3.connect('assets/quiz.db')
            cursor = conn.cursor()
            if selected_quiz == "Все викторины":
                cursor.execute("SELECT * FROM UserResult")
            else:
                cursor.execute("SELECT * FROM UserResult WHERE TestName=?", (selected_quiz,))
            rows = cursor.fetchall()
            conn.close()
        else:
            # Получаем данные из таблицы UserResult для указанного имени пользователя и выбранного названия викторины
            conn = sqlite3.connect('assets/quiz.db')
            cursor = conn.cursor()
            if selected_quiz == "Все викторины":
                cursor.execute("SELECT * FROM UserResult WHERE UserName=?", (search_name,))
            else:
                cursor.execute("SELECT * FROM UserResult WHERE UserName=? AND TestName=?", (search_name, selected_quiz))
            rows = cursor.fetchall()
            conn.close()

        # Вставляем найденные данные в Treeview
        for row in rows:
            tree.insert("", tk.END, values=(row[0], row[1], row[2], row[3]))

    def delete_records():
        selected_items = tree.selection()  # Получаем выбранные элементы в Treeview
        for item in selected_items:
            result_id = tree.item(item)['values'][0]  # Получаем ResultID выбранной записи
            tree.delete(item)  # Удаляем выбранный элемент из Treeview

            # Удаление записи из базы данных
            conn = sqlite3.connect('assets/quiz.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM UserResult WHERE ResultID=?", (result_id,))
            conn.commit()
            conn.close()

    def create_excel():
        workbook = Workbook()
        worksheet = workbook.active

        # Заголовки столбцов
        columns = ["ResultID", "UserName", "Score", "TestName"]
        for col_num, column_title in enumerate(columns, 1):
            worksheet.cell(row=1, column=col_num, value=column_title)

        # Данные из Treeview
        for row_num, item in enumerate(tree.get_children(), 2):
            values = tree.item(item)['values']
            for col_num, value in enumerate(values, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Сохранение файла
        file_name = "quiz_results.xlsx"
        workbook.save(file_name)

        # Открытие Excel-таблицы с использованием программы Excel
        os.system(f'start excel.exe "{file_name}"')

    def open_data_grid():
        create_data_grid()

    admin = tk.Tk()
    admin.title("Admin Page")
    admin.geometry("1200x600")
    window_width = 1200
    window_height = 600

    screen_width = admin.winfo_screenwidth()
    screen_height = admin.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    admin.geometry(f"{window_width}x{window_height}+{x}+{y}")
    # Creating search entry and button
    search_frame = tk.Frame(admin)
    search_frame.pack(side="top")

    search_label = tk.Label(search_frame, text="Введите имя пользователя:")
    search_label.pack(side="left")

    data_grid_button = tk.Button(search_frame, text="Обучающиеся", command=open_data_grid)
    data_grid_button.pack(side="right")

    search_entry = tk.Entry(search_frame)
    search_entry.pack(side="left")

    search_button = tk.Button(search_frame, text="Поиск", command=search_records)
    search_button.pack(side="left")

    # Creating ComboBox for quiz selection
    quiz_frame = tk.Frame(admin)
    quiz_frame.pack(side="top")

    quiz_label = tk.Label(quiz_frame, text="Выберите викторину:")
    quiz_label.pack(side="left")

    quiz_combobox = ttk.Combobox(quiz_frame, values=["Все викторины", "Работа с функциями и модулями", "Основы языка Python", "Работа с данными в Python"])
    quiz_combobox.current(0)  # Устанавливаем выбор по умолчанию на "Все викторины"
    quiz_combobox.pack(side="left")

    create_excel_button = tk.Button(admin, text="Создать Excel", command=create_excel)
    create_excel_button.pack(side="top")

    delete_button = tk.Button(admin, text="Удалить", command=delete_records)
    delete_button.pack(side="top")

    # Creating Treeview
    tree = ttk.Treeview(admin, show="headings")
    tree["columns"] = ("ResultID", "UserName", "Score", "TestName")
    tree["displaycolumns"] = ("ResultID", "UserName", "Score", "TestName")

    # Setting column headings
    tree.heading("ResultID", text="Номер результата")
    tree.heading("UserName", text="Имя пользователя")
    tree.heading("Score", text="Оценка")
    tree.heading("TestName", text="Название викторины")

    # Getting data from the UserResult table
    conn = sqlite3.connect('assets/quiz.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM UserResult")
    rows = cursor.fetchall()

    # Inserting data into Treeview
    for row in rows:
        tree.insert("", tk.END, values=(row[0], row[1], row[2], row[3]))  # Using column indices

    conn.close()

    # Placing the Treeview on the page
    tree.pack(fill=tk.BOTH, expand=True)

    admin.mainloop()


def signUpPage():
    root.destroy()
    global sup
    sup = Tk()
    # Центрирование окна
    window_width = 920
    window_height = 640

    screen_width = sup.winfo_screenwidth()
    screen_height = sup.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    sup.geometry(f"{window_width}x{window_height}+{x}+{y}")
    fname = StringVar()
    uname = StringVar()
    passW = StringVar()
    country = StringVar()
    show_password = BooleanVar()

    sup_canvas = Canvas(sup, width=920, height=640, bg="pink")
    sup_canvas.pack()

    sup_frame = Frame(sup_canvas, bg="white")
    sup_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    heading = Label(sup_frame, text="Регистрация", fg="black", bg="white")
    heading.config(font=('calibri 40'))
    heading.place(relx=0.30, rely=0.1)

    # full name
    flabel = Label(sup_frame, text="ФИО", fg='black', bg='white')
    flabel.place(relx=0.24, rely=0.4)
    fname = Entry(sup_frame, bg='#d3d3d3', fg='black', textvariable=fname)
    fname.config(width=42)
    fname.place(relx=0.31, rely=0.4)

    # username
    ulabel = Label(sup_frame, text="Имя пользователя", fg='black', bg='white')
    ulabel.place(relx=0.16, rely=0.5)
    user = Entry(sup_frame, bg='#d3d3d3', fg='black', textvariable=uname)
    user.config(width=42)
    user.place(relx=0.31, rely=0.5)

    # password
    plabel = Label(sup_frame, text="Пароль", fg='black', bg='white')
    plabel.place(relx=0.215, rely=0.6)
    pas = Entry(sup_frame, bg='#d3d3d3', fg='black', show="*", textvariable=passW)
    pas.config(width=42)
    pas.place(relx=0.31, rely=0.6)

    def toggle_password_visibility():
        if show_password.get():
            pas.config(show="")
        else:
            pas.config(show="*")



    # "Show Password" checkbox
    show_password_checkbox = Checkbutton(sup_frame, text="Показать пароль", variable=show_password,
                                         command=toggle_password_visibility)
    show_password_checkbox.place(relx=0.41, rely=0.75)
    # Group name options
    group_name_options = ["ИС-22", "ИС-21", "205", "195", "185", "175"]
    selected_group_name = StringVar(value=group_name_options[0])

    # Group name dropdown list
    group_name_label = Label(sup_frame, text="Номер группы", fg='black', bg='white')
    group_name_label.place(relx=0.15, rely=0.65)
    group_name_dropdown = OptionMenu(sup_frame, selected_group_name, *group_name_options)
    group_name_dropdown.config(width=8)
    group_name_dropdown.place(relx=0.41, rely=0.65)

    def addUserToDataBase():
        fullname = fname.get()
        username = user.get()
        password = pas.get()
        group_name = selected_group_name.get()

        conn = sqlite3.connect('assets/quiz.db')
        create = conn.cursor()
        create.execute('CREATE TABLE IF NOT EXISTS userSignUp(FULLNAME text, USERNAME text,PASSWORD text,GROUPNAME text)')
        create.execute("INSERT INTO userSignUp VALUES (?,?,?,?)", (fullname, username, password, group_name))
        conn.commit()
        create.execute('SELECT * FROM userSignUp')
        z = create.fetchall()
        print(z)
        #        L2.config(text="Username is "+z[0][0]+"\nPassword is "+z[-1][1])
        conn.close()
        loginPage(z)

    def gotoLogin():
        conn = sqlite3.connect('assets/quiz.db')
        create = conn.cursor()
        conn.commit()
        create.execute('SELECT * FROM userSignUp')
        z = create.fetchall()
        loginPage(z)

    # signup BUTTON
    sp = Button(sup_frame, text='Зарегистрироваться', padx=4, pady=4, width=4, command=addUserToDataBase, bg='green')
    sp.configure(width=15, height=1, activebackground="#33B5E5", relief=FLAT)
    sp.place(relx=0.41, rely=0.82)

    log = Button(sup_frame, text='Уже есть учётная запись?', padx=5, pady=5, width=5, command=gotoLogin, bg="white",
                 fg='blue')
    log.configure(width=18, height=1, activebackground="#33B5E5", relief=FLAT)
    log.place(relx=0.39, rely=0.9)

    sup.mainloop()


def menu():
    login.destroy()
    global menu
    menu = Tk()
    # Центрирование окна
    window_width = 720
    window_height = 440

    screen_width = menu.winfo_screenwidth()
    screen_height = menu.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    menu.geometry(f"{window_width}x{window_height}+{x}+{y}")
    menu_canvas = Canvas(menu, width=720, height=440, bg="pink")
    menu_canvas.pack()

    menu_frame = Frame(menu_canvas, bg="white")
    menu_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    wel = Label(menu_canvas, text=' П р о й д и т е  о п р о с ', fg="white", bg="pink")
    wel.config(font=('Broadway 22'))
    wel.place(relx=0.25, rely=0.01)

    level = Label(menu_frame, text='Выберите тему викторины', bg="pink", font="calibri 18")
    level.place(relx=0.20, rely=0.3)

    var = IntVar()
    easyR = Radiobutton(menu_frame, text='Основы языка Python', bg="white", font="calibri 16", value=1, variable=var)
    easyR.place(relx=0.25, rely=0.4)

    mediumR = Radiobutton(menu_frame, text='Работа с функциями и модулями в Python', bg="white", font="calibri 16",
                          value=2, variable=var)
    mediumR.place(relx=0.25, rely=0.5)

    hardR = Radiobutton(menu_frame, text='Работа с данными в Python', bg="white", font="calibri 16", value=3,
                        variable=var)
    hardR.place(relx=0.25, rely=0.6)

    def navigate():
        x = var.get()
        print(x)
        if x == 1:
            easy()
        elif x == 2:
            medium()
        elif x == 3:
            difficult()
        else:
            pass

    letsgo = Button(menu_frame, text="Let's Go", bg="white", font="calibri 12", command=navigate)
    letsgo.place(relx=0.25, rely=0.8)

    def logout():
        menu.destroy()
        start()


    logout_button = Button(menu_frame, text="Выйти", bg="white", font="calibri 12", command=logout)
    logout_button.place(relx=0.5, rely=0.8)

    menu.mainloop()

def easy():
    global e
    e = Tk()
    global testName
    global score

    score = 4  # Инициализация переменной score
    testName = "Основы языка Python"
    window_width = 820
    window_height = 460
    screen_width = e.winfo_screenwidth()
    screen_height = e.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    e.geometry(f"{window_width}x{window_height}+{x}+{y}")  # Установка размеров и позиции окна

    easy_canvas = Canvas(e, width=window_width, height=window_height, bg="#101357")
    easy_canvas.pack()

    easy_frame = Frame(easy_canvas, bg="white")
    easy_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    def countDown():
        check = 0
        for k in range(10, 0, -1):
            if k == 1:
                check = -1
            timer.configure(text=k)
            easy_frame.update()
            time.sleep(1)

        timer.configure(text="Times up!")
        if check == -1:
            return -1
        else:
            return 0

    easyQ = [
        [
            "Какая команда используется для вывода текста на экран в Python?",
            "print()",
            "input()",
            "scan()",
            "read()"
        ],
        [
            "Что такое переменная в Python и как ее объявить?",
            "Это функция для хранения данных. Объявляется с помощью ключевого слова variable.",
            "Это команда для выполнения определенного действия. Объявляется с помощью ключевого слова var.",
            "Это имя, используемое для обозначения значения. Объявляется с помощью оператора =.",
            "Это способ обращения к элементам списка. Объявляется с помощью оператора +."
        ],
        [
            "Какой символ используется для обозначения однострочного комментария в Python?",
            "#",
            "//",
            "/*",
            "~"
        ],
        [
            "Выберите код, который считает сумму двух чисел и выводит результат",
            "sum(2, 3)",
            "print(sum(2, 3))",
            "add(2, 3)",
            "print(add(2, 3))"
        ],
        [
            "Что такое условный оператор if-else и как он используется в Python?",
            "Это оператор для создания цикла. Используется для повторения определенного блока кода.",
            "Это оператор для проверки условия. Если условие истинно, выполняется один блок кода, если ложно - другой.",
            "Это оператор для объявления функции. Используется для выполнения определенного действия.",
            "Это оператор для работы с текстовыми данными. Используется для поиска и замены символов."
        ]
    ]
    answer = [
        "print()",
        "Это функция для хранения данных. Объявляется с помощью ключевого слова variable.",
        "#",
        "print(sum(2, 3))",
        "Это оператор для проверки условия. Если условие истинно, выполняется один блок кода, если ложно - другой."
    ]
    li = list(range(len(easyQ)))

    x = random.choice(li)

    ques = Label(easy_frame, text=easyQ[x][0], font="calibri 12", bg="white")
    ques.place(relx=0.5, rely=0.2, anchor=CENTER)

    var = StringVar()

    a = Radiobutton(easy_frame, text=easyQ[x][1], font="calibri 10", value=easyQ[x][1], variable=var, bg="white")
    a.place(relx=0.5, rely=0.42, anchor=CENTER)

    b = Radiobutton(easy_frame, text=easyQ[x][2], font="calibri 10", value=easyQ[x][2], variable=var, bg="white")
    b.place(relx=0.5, rely=0.52, anchor=CENTER)

    c = Radiobutton(easy_frame, text=easyQ[x][3], font="calibri 10", value=easyQ[x][3], variable=var, bg="white")
    c.place(relx=0.5, rely=0.62, anchor=CENTER)

    d = Radiobutton(easy_frame, text=easyQ[x][4], font="calibri 10", value=easyQ[x][4], variable=var, bg="white")
    d.place(relx=0.5, rely=0.72, anchor=CENTER)

    li.remove(x)

    timer = Label(e)
    timer.place(relx=0.8, rely=0.82, anchor=CENTER)

    def display():
        if len(li) == 1:
            e.destroy()
            showMark(score, testName)
        if len(li) == 2:
            nextQuestion.configure(text='End', command=calc)

        if li:
            x = random.choice(li)
            ques.configure(text=easyQ[x][0])

            a.configure(text=easyQ[x][1], value=easyQ[x][1])
            b.configure(text=easyQ[x][2], value=easyQ[x][2])
            c.configure(text=easyQ[x][3], value=easyQ[x][3])
            d.configure(text=easyQ[x][4], value=easyQ[x][4])

            li.remove(x)
            y = countDown()
            if y == -1:
                display()

    def calc():
        selected_answer = var.get()
        if selected_answer == answer[x]:
            global score
            score += 1  # Увеличение оценки score на 1
        display()

    submit = Button(easy_frame, command=calc, text="Submit")
    submit.place(relx=0.5, rely=0.82, anchor=CENTER)

    nextQuestion = Button(easy_frame, command=display, text="Next")
    nextQuestion.place(relx=0.87, rely=0.82, anchor=CENTER)

    y = countDown()
    if y == -1:
        display()
    e.mainloop()


def medium():

    testName = "Работа с функциями и модулями"

    global m
    m = Tk()
    window_width = 720
    window_height = 440

    screen_width = m.winfo_screenwidth()
    screen_height = m.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    m.geometry(f"{window_width}x{window_height}+{x}+{y}")
    med_canvas = Canvas(m, width=720, height=440, bg="#101357")
    med_canvas.pack()

    med_frame = Frame(med_canvas, bg="white")
    med_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    global score
    score = 0

    mediumQ = [
        [
            "Объявите функцию calc",
            "def calc()",
        ],
        [
            "Напишите код для импорта модуля tkinter",
            "import tkinter",
        ],
        [
            "Напишите код, который выводит строку Hello, World! на экран в одну строку",
            "print('Hello, World!')",
        ],
        [
            "Напишите код, который выводит числа от 1 до 10 в одну строку",
            "print(*range(1, 11))",
        ],
        [
            "Напишите код, который выводит квадраты чисел от 1 до 5 в одну строку",
            "print(*[x**2 for x in range(1, 6)])",
        ],
    ]
    answer = [
        "def calc()",
        "import tkinter",
        "print('Hello, World!')",
        "print(*range(1, 11))",
        "print(*[x**2 for x in range(1, 6)])",
    ]

    li = [0, 1, 2, 3, 4]
    random.shuffle(li)
    question_index = 0

    ques = Label(med_frame, text=mediumQ[li[question_index]][0], font="calibri 12", bg="white")
    ques.pack(pady=10, padx=10, anchor=CENTER)

    entry = Entry(med_frame, width=30, font="calibri 10")
    entry.pack(pady=10, padx=10, anchor=CENTER)

    def check_answer():
        user_answer = entry.get().strip()
        if user_answer == answer[li[question_index]]:
            global score
            score += 1

            next_question()


    def next_question():
        nonlocal question_index
        question_index += 1
        if question_index < len(li):
            ques.configure(text=mediumQ[li[question_index]][0])
            entry.delete(0, END)
        else:
            m.destroy()
            showMark(score, testName)


    submit = Button(med_frame, command=check_answer, text="Submit")
    submit.pack(pady=10, padx=10, anchor=CENTER)

    nextQuestion = Button(med_frame, command=next_question, text="Next")
    nextQuestion.pack(pady=10, padx=10, anchor=CENTER)

    m.mainloop()


def difficult():
    testName = "Работа с данными в Python"



    def check_order():

        if user_order == correct_order:
            success_label.configure(text="Порядок верный! Успех!", fg="green")
            score_label.configure(text="Оценка: 5 баллов")
            global score
            score = 5
            showMark(score, testName)
        else:
            success_label.configure(text="Неправильный порядок. Попробуйте еще раз.", fg="red")

    def move_up():
        selected_index = code_listbox.curselection()
        if selected_index:
            index = selected_index[0]
            if index > 0:
                item = user_order.pop(index)
                user_order.insert(index - 1, item)
                update_code_listbox()
                code_listbox.selection_set(index - 1)

    def move_down():
        selected_index = code_listbox.curselection()
        if selected_index:
            index = selected_index[0]
            if index < len(user_order) - 1:
                item = user_order.pop(index)
                user_order.insert(index + 1, item)
                update_code_listbox()
                code_listbox.selection_set(index + 1)

    def update_code_listbox():
        code_listbox.delete(0, tk.END)
        for code in user_order:
            code_listbox.insert(tk.END, code)

    # Правильный порядок строк кода
    correct_order = [
        "def calc(a, b, operator):",
        "    if operator == '+':",
        "        return a + b",
        "    elif operator == '-':",
        "        return a - b",
        "    elif operator == '*':",
        "        return a * b",
        "    elif operator == '/':",
        "        if b != 0:",
        "            return a / b",
        "        else:",
        "            return 'Error: division by zero'",
        "",
        "result = calc(5, 2, '+')",
        "print('Result:', result)"
    ]

    # Случайное перемешивание строк кода для пользователя
    user_order = correct_order.copy()
    random.shuffle(user_order)

    # Инициализация окна игры
    h = tk.Tk()
    h.geometry("450x450")
    h.title("Разместите код в нужном порядке.Порядок: +-*/")
    # Определение размеров окна
    window_width = 450
    window_height = 450

    # Получение размеров экрана
    screen_width = h.winfo_screenwidth()
    screen_height = h.winfo_screenheight()

    # Вычисление координат центра экрана
    x = int((screen_width / 2) - (window_width / 2))
    y = int((screen_height / 2) - (window_height / 2))

    # Установка положения окна
    h.geometry(f"{window_width}x{window_height}+{x}+{y}")
    # Создание и расположение виджетов
    code_listbox = tk.Listbox(h, width=50, height=15, selectmode=tk.SINGLE)
    code_listbox.pack()

    success_label = tk.Label(h, text="", font="Arial 12 bold")
    success_label.pack()

    score_label = tk.Label(h, text="Оценка: 0 баллов", font="Arial 12 bold")
    score_label.pack()

    check_button = tk.Button(h, text="Проверить порядок", command=check_order)
    check_button.pack()

    update_code_listbox()

    h.bind("<Up>", lambda event: move_up())
    h.bind("<Down>", lambda event: move_down())

    h.mainloop()


def showMark(mark, testName):
    sh = tk.Toplevel()
    sh.title("Результаты викторины")

    window_width = 720
    window_height = 440

    screen_width = sh.winfo_screenwidth()
    screen_height = sh.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    sh.geometry(f"{window_width}x{window_height}+{x}+{y}")
    show_canvas = tk.Canvas(sh, width=720, height=440, bg="#101357")
    show_canvas.pack()

    show_frame = tk.Frame(show_canvas, bg="white")
    show_frame.place(relwidth=0.8, relheight=0.8, relx=0.1, rely=0.1)

    st = f"{global_username}, Ваша оценка {mark}"

    title_label = tk.Label(show_frame, text="Результаты викторины", font=("Arial", 20), bg="white")
    title_label.pack(pady=20)

    score_label = tk.Label(show_frame, text=st, font=("Arial", 16), bg="white")
    score_label.pack(pady=20)

    image_path = "succecimage.png" if mark >= 4 else "unsucces.png"
    image = tk.PhotoImage(file=image_path)

    image_label = tk.Label(show_frame, image=image, bg="white")
    image_label.image = image  # Сохраняем ссылку на изображение, чтобы избежать ошибки
    image_label.pack(pady=20)

    conn = sqlite3.connect('assets/quiz.db')
    cursor = conn.cursor()

    cursor.execute('''CREATE TABLE IF NOT EXISTS UserResult
                    (UserName TEXT, Score INT, TestName)''')

    cursor.execute("INSERT INTO UserResult (UserName, Score, TestName) VALUES (?, ?, ?)",
                   (global_username, mark, testName))

    conn.commit()
    conn.close()

    sh.mainloop()
def create_data_grid():
    # Создание окна
    window = tk.Tk()
    window.title("Data Grid")

    # Увеличение окна
    window_width = 800
    window_height = 600

    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # Создание DataGrid
    data_grid = ttk.Treeview(window)
    data_grid["columns"] = ("fullname", "username", "password", "groupname")
    data_grid.column("#0", width=0, stretch=tk.NO)
    data_grid.column("fullname", anchor=tk.W, width=100)
    data_grid.column("username", anchor=tk.W, width=100)
    data_grid.column("password", anchor=tk.W, width=100)
    data_grid.column("groupname", anchor=tk.W, width=100)

    data_grid.heading("#0", text="")
    data_grid.heading("fullname", text="Full Name")
    data_grid.heading("username", text="Username")
    data_grid.heading("password", text="Password")
    data_grid.heading("groupname", text="Group Name")

    data_grid.pack(fill=tk.BOTH, expand=True)

    # Подключение к базе данных
    conn = sqlite3.connect("assets/quiz.db")
    cursor = conn.cursor()

    def fetch_data():
        # Очистка DataGrid перед загрузкой новых данных
        data_grid.delete(*data_grid.get_children())

        # Получение данных из таблицы userSignUp
        cursor.execute("SELECT FULLNAME, USERNAME, PASSWORD, GROUPNAME FROM userSignUp")
        rows = cursor.fetchall()

        # Заполнение DataGrid данными из таблицы
        for row in rows:
            data_grid.insert("", tk.END, values=row)

    def add_data():
        def save_data():
            fullname = fullname_entry.get()
            username = username_entry.get()
            password = password_entry.get()
            groupname = groupname_entry.get()

            # Вставка новой записи в таблицу userSignUp
            cursor.execute("INSERT INTO userSignUp (FULLNAME, USERNAME, PASSWORD, GROUPNAME) VALUES (?, ?, ?, ?)",
                           (fullname, username, password, groupname))
            conn.commit()

            # Обновление отображения DataGrid
            fetch_data()

            # Закрытие окна добавления данных
            add_window.destroy()

        # Окно для добавления данных
        add_window = tk.Toplevel(window)
        add_window.title("Добавить данные")

        # Увеличение окна
        window_width = 400
        window_height = 300

        screen_width = add_window.winfo_screenwidth()
        screen_height = add_window.winfo_screenheight()

        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        add_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

        fullname_label = tk.Label(add_window, text="ФИО:")
        fullname_label.pack()
        fullname_entry = tk.Entry(add_window)
        fullname_entry.pack()

        username_label = tk.Label(add_window, text="Имя пользователя:")
        username_label.pack()
        username_entry = tk.Entry(add_window)
        username_entry.pack()

        password_label = tk.Label(add_window, text="Пароль:")
        password_label.pack()
        password_entry = tk.Entry(add_window)
        password_entry.pack()

        groupname_label = tk.Label(add_window, text="Название группы:")
        groupname_label.pack()
        groupname_entry = tk.Entry(add_window)
        groupname_entry.pack()

        save_button = tk.Button(add_window, text="Сохранить", command=save_data)
        save_button.pack()

    def update_data():
        selected_item = data_grid.selection()
        if not selected_item:
            messagebox.showwarning("Внимание", "Не выбраны данные")
            return

        # Получение данных выбранной записи
        values = data_grid.item(selected_item)["values"]
        fullname, username, password, groupname = values

        def save_changes():
            new_fullname = fullname_entry.get()
            new_username = username_entry.get()
            new_password = password_entry.get()
            new_groupname = groupname_entry.get()

            # Обновление выбранной записи в таблице userSignUp
            cursor.execute("UPDATE userSignUp SET FULLNAME=?, USERNAME=?, PASSWORD=?, GROUPNAME=? WHERE USERNAME=?",
                           (new_fullname, new_username, new_password, new_groupname, username))
            conn.commit()

            # Обновление отображения DataGrid
            fetch_data()

            # Закрытие окна редактирования данных
            edit_window.destroy()

        # Окно для редактирования данных
        edit_window = tk.Toplevel(window)
        edit_window.title("Редактировать данные")

        # Увеличение окна
        window_width = 400
        window_height = 300

        screen_width = edit_window.winfo_screenwidth()
        screen_height = edit_window.winfo_screenheight()

        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        edit_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

        fullname_label = tk.Label(edit_window, text="ФИО:")
        fullname_label.pack()
        fullname_entry = tk.Entry(edit_window)
        fullname_entry.pack()
        fullname_entry.insert(tk.END, fullname)

        username_label = tk.Label(edit_window, text="Имя пользователя:")
        username_label.pack()
        username_entry = tk.Entry(edit_window)
        username_entry.pack()
        username_entry.insert(tk.END, username)
        username_entry.config(state="readonly")

        password_label = tk.Label(edit_window, text="Пароль:")
        password_label.pack()
        password_entry = tk.Entry(edit_window)
        password_entry.pack()
        password_entry.insert(tk.END, password)

        groupname_label = tk.Label(edit_window, text="Название группы:")
        groupname_label.pack()
        groupname_entry = tk.Entry(edit_window)
        groupname_entry.pack()
        groupname_entry.insert(tk.END, groupname)

        save_button = tk.Button(edit_window, text="Сохранить изменения", command=save_changes)
        save_button.pack()

    def delete_data():
        selected_item = data_grid.selection()
        if not selected_item:
            messagebox.showwarning("Предупреждение!", "Данные не выбраны.")
            return

        confirmation = messagebox.askyesno("Подтверждение", "Вы действительно желаете удалить выбранную запись?")
        if confirmation:
            # Получение данных выбранной записи
            values = data_grid.item(selected_item)["values"]
            username = values[1]

            # Удаление выбранной записи из таблицы userSignUp
            cursor.execute("DELETE FROM userSignUp WHERE USERNAME=?", (username,))
            conn.commit()

            # Обновление отображения DataGrid
            fetch_data()

    # Кнопки для CRUD операций
    buttons_frame = tk.Frame(window)
    buttons_frame.pack(pady=10)

    add_button = tk.Button(buttons_frame, text="Добавить", command=add_data)
    add_button.pack(side=tk.LEFT)

    update_button = tk.Button(buttons_frame, text="Обновить", command=update_data)
    update_button.pack(side=tk.LEFT)

    delete_button = tk.Button(buttons_frame, text="Удалить", command=delete_data)
    delete_button.pack(side=tk.LEFT)

    # Загрузка и отображение данных в DataGrid
    fetch_data()

    # Запуск главного цикла окна
    window.mainloop()

# Вызов функции для создания страницы с DataGrid


def start():
    global root
    root = Tk()
    # Размеры окна
    window_width = 1280
    window_height = 744

    # Получение размеров экрана
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Вычисление позиции окна для центрирования
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # Установка размеров и позиции окна
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    canvas = Canvas(root, width=1280, height=700)
    canvas.grid(column=0, row=1)
    img = PhotoImage(file="background.png")
    canvas.create_image(50, 10, image=img, anchor=NW)

    button = Button(root, text='Начать', command=signUpPage)
    button.configure(width=102, height=2, activebackground="#33B5E5", bg='green', relief=RAISED)
    button.grid(column=0, row=2)

    root.mainloop()


if __name__ == '__main__':
    start()
